import sys
from openpyxl import load_workbook

class Process:
    def __init__(self, pid, arrival_time, instruction_load, priority):
        self.pid = pid
        self.arrival_time = arrival_time
        self.instructions = instruction_load
        self.initial_instructions = self.instructions
        self.priority = priority

class Scheduler:
    def __init__(self, filename="cpu-scheduling.xlsx"):
        self.filename = filename
        self.processes = []

    def load_processes(self):
        work_book = load_workbook(filename=self.filename)
        for line in work_book.active.iter_rows():
            if isinstance(line[0].value, str):
                continue
            my_process = Process(
                pid=int(line[0].value),
                arrival_time=int(line[1].value),
                instruction_load=int(line[2].value),
                priority=int(line[3].value)
            )
            self.processes.append(my_process)

    def run_fcfs(self):
        time_unit = 0
        last_process = None
        while time_unit < 400:
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            first_process = self.get_first_process(current_processes)
            is_switching_context = False
            if first_process != last_process:
                if last_process is not None:
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = first_process
            if first_process is not None and is_switching_context is False:
                self.process_and_output(time_unit, first_process, current_processes)

    def run_round_robin(self):
        time_unit = 0
        last_process = None
        while time_unit < 400:
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            first_process = self.get_first_process_rr(current_processes)
            is_switching_context = False
            if first_process is not None and first_process.bulk_runs >= 4:
                first_process.bulk_runs = 0
                self.processes.remove(first_process)
                self.processes.append(first_process)
                if len(current_processes) > 0:
                    first_process = current_processes.pop(0)
                else:
                    first_process = None
            if first_process != last_process:
                if last_process is not None:
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = first_process
            if first_process is not None and is_switching_context is False:
                first_process.bulk_runs += 1
                self.process_and_output(time_unit, first_process, current_processes)

    def run_sjf(self):
        last_process = None
        time_unit = 0
        while time_unit < 400:
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            shortest_process, waiting_processes = self.get_shortest_process(current_processes)
            is_switching_context = False
            if shortest_process != last_process:
                if last_process is not None:
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = shortest_process
            if shortest_process is not None and is_switching_context is False:
                self.process_and_output(time_unit, shortest_process, waiting_processes)

    def run_priority_queue(self):
        last_process = None
        time_unit = 0
        while time_unit < 400:
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            max_process, waiting_processes = self.get_process_with_highest_priority(current_processes)
            is_switching_context = False
            if max_process != last_process:
                if last_process is not None:
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = max_process
            if max_process is not None and is_switching_context is False:
                self.process_and_output(time_unit, max_process, waiting_processes)

    def get_processes_for_time_unit(self, time_unit):
        current_processes = []
        for my_process in self.processes:
            if my_process.arrival_time <= time_unit and my_process.instructions > 0:
                current_processes.append(my_process)
        return current_processes

    def get_first_process(self, current_processes):
        if len(current_processes) > 0:
            return current_processes.pop(0)
        return None

    def get_first_process_rr(self, current_processes):
        if len(current_processes) > 0:
            return current_processes.pop(0)
        return None

    def get_shortest_process(self, current_processes):
        shortest_instructions = 1000
        shortest_process = None
        waiting_processes = []
        for my_process in current_processes:
            if my_process.instructions < shortest_instructions:
                if shortest_process is not None:
                    waiting_processes.append(shortest_process)
                shortest_process = my_process
                shortest_instructions = my_process.instructions
            else:
                waiting_processes.append(my_process)
        return shortest_process, waiting_processes

    def get_process_with_highest_priority(self, current_processes):
        max_priority = 0
        max_process = None
        waiting_processes = []
        for my_process in current_processes:
            if my_process.priority > max_priority:
                if max_process is not None:
                    waiting_processes.append(max_process)
                max_process = my_process
                max_priority = my_process.priority
            else:
                waiting_processes.append(my_process)
        return max_process, waiting_processes

    @staticmethod
    def get_waiting_processes(max_process, current_processes):
        waiting_processes = []
        for my_process in current_processes:
            if my_process.pid != max_process.pid:
                waiting_processes.append(my_process)
        return waiting_processes

    @staticmethod
    def process_and_output(time_unit, max_process, waiting_process):
        max_process.instructions -= 1
        print(f"Time Unit: {time_unit}, PID {max_process.pid} executing, Instructions left: {max_process.instructions}")
        waiting_text = ""
        i = 0
        for my_process in waiting_process:
            i += 1
            if i % 6 == 0:
                waiting_text = f"{waiting_text}\n"
            waiting_time = time_unit - my_process.arrival_time - (my_process.initial_instructions - my_process.instructions) + 1
            waiting_text = f"{waiting_text}PID {my_process.pid} waiting time={waiting_time}. "
        if waiting_text != "":
            print(waiting_text)

def main():
    if len(sys.argv) != 2:
        print("Usage: python script.py <algorithm>")
        print("Available algorithms: fcfs, roundrobin, sjf, priorityqueue")
        sys.exit(1)

    algorithm = sys.argv[1].lower()
    my_scheduler = Scheduler()

    if algorithm == "fcfs":
        my_scheduler.load_processes()
        my_scheduler.run_fcfs()
    elif algorithm == "roundrobin":
        my_scheduler.load_processes()
        my_scheduler.run_round_robin()
    elif algorithm == "sjf":
        my_scheduler.load_processes()
        my_scheduler.run_sjf()
    elif  algorithm == "priorityqueue":
        my_scheduler.load_processes()
        my_scheduler.run_priority_queue()
    else:
        print("Invalid algorithm. Available options: fcfs, roundrobin, sjf, priorityqueue")
        sys.exit(1)

if __name__ == "__main__":
    main()

