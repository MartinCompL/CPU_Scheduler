# Import necessary module to work with Excel files
from openpyxl import load_workbook

# Load the Excel workbook
wb = load_workbook(filename="cpu-scheduling.xlsx")
sheet = wb.active

# Constructor for the process
class Process:
    # Initialize process attributes
    def __init__(self, pid, arrival_time, instruction_load, priority):
        self.pid = pid
        self.arrival_time = arrival_time
        self.instructions = instruction_load
        self.initial_instructions = self.instructions
        self.priority = priority
        self.bulk_runs = 0

# Blueprint to sort and trigger the output
class Scheduler:
    # Initialize Scheduler with optional filename parameter
    def __init__(self, filename="cpu-scheduling.xlsx"):
        self.filename = filename
        self.processes = []

    # Static method to sort processes by arrival time
    @staticmethod
    def sort_by_arrival_time(my_process):
        return my_process.arrival_time

    # Load processes from Excel workbook and append them to the list
    def load_processes(self):
        work_book = load_workbook(filename="cpu-scheduling.xlsx")
        # Iterate through each row in the workbook
        for line in work_book.active.iter_rows():
            # Skip header row and continue with data rows
            if isinstance(line[0].value, str):
                continue
            # Create a Process instance for each row and append to the processes list
            my_process = Process(
                pid=int(line[0].value),
                arrival_time=int(line[1].value),
                instruction_load=int(line[2].value),
                priority=int(line[3].value)
            )
            self.processes.append(my_process)
            # Sort processes based on arrival time
            self.processes.sort(key=Scheduler.sort_by_arrival_time)

    # Round Robin algorithm for process execution
    def run(self):
        time_unit = 0
        last_process = None
        while time_unit < 400:
            first_process = None
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            # Check if there are processes available for execution
            if len(current_processes) > 0:
                # Pop the first process from the list
                first_process = current_processes.pop(0)
            # Check if the process has completed its time quantum
            if first_process is not None and first_process.bulk_runs >= 4:
                # Reset bulk_runs and move the process to the end of the list
                first_process.bulk_runs = 0
                self.processes.remove(first_process)
                self.processes.append(first_process)
                # Check if there are more processes available for execution
                if len(current_processes) > 0:
                    first_process = current_processes.pop(0)
                else:
                    first_process = None
            is_switching_context = False
            # Check if there is a context switch between processes
            if first_process != last_process:
                if last_process is not None:
                    # Perform context switch if the last process is not done
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = first_process
            # Execute the process if it is not a context switch
            if first_process is not None and is_switching_context is False:
                first_process.bulk_runs += 1
                self.process_and_output(time_unit, first_process, current_processes)

    # Get the processes ready to run for the given time unit
    def get_processes_for_time_unit(self, time_unit):
        current_processes = []
        # Iterate through each process to check if it is ready for execution
        for my_process in self.processes:
            if my_process.arrival_time <= time_unit and my_process.instructions > 0:
                current_processes.append(my_process)
        return current_processes

    # Static method to get the shortest process and waiting processes
    @staticmethod
    def get_shortest_process(current_processes):
        shortest_instructions = 1000
        shortest_process = None
        waiting_processes = []
        # Iterate through each process to find the shortest one
        for my_process in current_processes:
            if my_process.instructions < shortest_instructions:
                if shortest_process is not None:
                    waiting_processes.append(shortest_process)
                shortest_process = my_process
                shortest_instructions = my_process.instructions
            else:
                waiting_processes.append(my_process)
        return shortest_process, waiting_processes

    # Static method to get processes waiting for execution
    @staticmethod
    def get_waiting_processes(max_process, current_processes):
        waiting_processes = []
        # Iterate through each process to find waiting processes
        for my_process in current_processes:
            if my_process.pid != max_process.pid:
                waiting_processes.append(my_process)
        return waiting_processes

    # Static method to process and output information in a formatted string
    @staticmethod
    def process_and_output(time_unit, shortest_process, waiting_process):
        shortest_process.instructions -= 1
        # Print the information about the process execution
        print(f"Time Unit {time_unit}: PID {shortest_process.pid} executing, Instructions left: {shortest_process.instructions}")
        waiting_text = ""
        i = 0
        # Format the waiting processes information
        for my_process in waiting_process:
            i += 1
            if i % 6 == 0:
                waiting_text = f"{waiting_text}\n"
            waiting_time = time_unit - my_process.arrival_time - (my_process.initial_instructions - my_process.instructions) + 1
            waiting_text = f"{waiting_text}PID {my_process.pid} waiting time={waiting_time}. "
        # Print the waiting processes information if it exists
        if waiting_text != "":
            print(waiting_text)

# Running modules
if __name__ == "__main__":
    # Create an instance of the Scheduler
    my_scheduler = Scheduler()
    # Load processes from the Excel workbook
    my_scheduler.load_processes()
    # Run the scheduler to execute processes
    my_scheduler.run()
