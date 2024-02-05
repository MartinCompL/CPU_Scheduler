# Import necessary module to work with Excel files
from openpyxl import load_workbook

# Load the Excel workbook
wb = load_workbook(filename="cpu-scheduling.xlsx")
sheet = wb.active

# Constructor for the process
class Process:
    # Initialize process attributes
    def __init__(self, pid, arrival_time, instruction_load, priority):
        self.pid = pid
        self.arrival_time = arrival_time
        self.instructions = instruction_load
        self.initial_instructions = self.instructions
        self.priority = priority

# Initiating the cpu-scheduling.xlsx file
class Scheduler:
    # Initialize Scheduler with optional filename parameter
    def __init__(self, filename="cpu-scheduling.xlsx"):
        self.filename = filename
        self.processes = []

    # Load processes from Excel workbook and append them to the list
    def load_processes(self):
        work_book = load_workbook(filename="cpu-scheduling.xlsx")
        # Iterate through each row in the workbook
        for line in work_book.active.iter_rows():
            # Skip header row and continue with data rows
            if isinstance(line[0].value, str):
                # 'continue' keyword is used to end the current iteration and continue to the next one
                continue
            # Create a Process instance for each row and append to the processes list
            my_process = Process(
                pid=int(line[0].value),
                arrival_time=int(line[1].value),
                instruction_load=int(line[2].value),
                priority=int(line[3].value)
            )
            self.processes.append(my_process)

    # Priority Queue algorithm for process execution
    def run(self):
        last_process = None
        time_unit = 0
        # Time unit cannot go above 400
        while time_unit < 400:
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            max_process, waiting_processes = self.get_process_with_highest_priority(current_processes)
            is_switching_context = False
            # Check if there is a context switch between processes
            if max_process != last_process:
                # If the last process is not done, it will context switch
                if last_process is not None:
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = max_process
            # Execute the process if it is not a context switch
            if max_process is not None and is_switching_context is False:
                self.process_and_output(time_unit, max_process, waiting_processes)

    # Get the processes ready to run for the given time unit
    def get_processes_for_time_unit(self, time_unit):
        current_processes = []
        # Iterate through each process to check if it is ready for execution
        for my_process in self.processes:
            if my_process.arrival_time <= time_unit and my_process.instructions > 0:
                current_processes.append(my_process)
        return current_processes

    # Static method to get the process with the highest priority and waiting processes
    @staticmethod
    def get_process_with_highest_priority(current_processes):
        max_priority = 0
        max_process = None
        waiting_processes = []
        # Iterate through each process to find the one with the highest priority
        for my_process in current_processes:
            if my_process.priority > max_priority:
                if max_process is not None:
                    waiting_processes.append(max_process)
                max_process = my_process
                max_priority = my_process.priority
            else:
                waiting_processes.append(my_process)
        return max_process, waiting_processes

    @staticmethod
    # Gets the waiting process and appends it
    def get_waiting_processes(max_process, current_processes):
        waiting_processes = []
        for my_process in current_processes:
            if my_process.pid != max_process.pid:
                waiting_processes.append(my_process)
        return waiting_processes

    @staticmethod
    # Output the information in a formatted string
    def process_and_output(time_unit, max_process, waiting_process):
        max_process.instructions -= 1
        # Print the information about the process execution
        print(f"Time Unit: {time_unit}, PID {max_process.pid} executing, Instructions left: {max_process.instructions}")
        waiting_text = ""
        i = 0
        # Format the waiting processes information
        for my_process in waiting_process:
            i += 1
            if i % 6 == 0:
                waiting_text = f"{waiting_text}\n"
            waiting_time = time_unit - my_process.arrival_time - (my_process.initial_instructions - my_process.instructions) + 1
            waiting_text = f"{waiting_text}PID {my_process.pid} waiting time={waiting_time}. "
        # Print the waiting processes information if it exists
        if waiting_text != "":
            print(waiting_text)

# Running modules
if __name__ == "__main__":
    # Create an instance of the Scheduler
    my_scheduler = Scheduler()
    # Load processes from the Excel workbook
    my_scheduler.load_processes()
    # Run the scheduler to execute processes
    my_scheduler.run()
