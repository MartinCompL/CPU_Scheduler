from openpyxl import load_workbook
 
wb = load_workbook(filename="cpu-scheduling.xlsx")
sheet = wb.active
 
#constructor for the process
class Process:
    def __init__(self, pid, arrival_time, instruction_load, priority):
        self.pid = pid
        self.arrival_time = arrival_time
        self.instructions = instruction_load
        self.initial_instructions = self.instructions
        self.priority = priority
 
#Initiating the cp-scheduling.xlsx file
class Scheduler:
    def __init__(self, filename="cpu-scheduling.xlsx"):
        self.filename = filename
        self.processes = []
#loads the process of values and appends
    def load_processes(self):
        work_book = load_workbook(filename="cpu-scheduling.xlsx")
        for line in work_book.active.iter_rows():
            if isinstance(line[0].value, str):
#continue keyword is used to ends the current interation and continue the next one
                continue
            my_process = Process(
                pid=int(line[0].value),
                arrival_time=int(line[1].value),
                instruction_load=int(line[2].value),
                priority=int(line[3].value)
            )
            self.processes.append(my_process)
 
#FCFS algorithm
    def run(self):
        last_process = None
        time_unit = 0
        #time_unit can not go above 400 time_units
        while time_unit < 400:
            first_process = None
            time_unit += 1
            current_processes = self.get_processes_for_time_unit(time_unit)
            if len(current_processes) > 0:
                first_process = current_processes.pop(0)
            is_switching_context = False
            if first_process != last_process:
                #if the last process is not done, it will context switch.
                if last_process is not None:
                    print(f"Time Unit {time_unit}: Context switch.")
                    is_switching_context = True
                last_process = first_process
            if first_process is not None and is_switching_context is False:
                self.process_and_output(time_unit, first_process, current_processes)
 
    #function to process the time units
    def get_processes_for_time_unit(self, time_unit):
        current_processes = []
        for my_process in self.processes:
            if my_process.arrival_time <= time_unit and my_process.instructions > 0:
                current_processes.append(my_process)
        return current_processes
    #staticmethod is a way to create a static function without creating an instance
    @staticmethod
    #output the information in a formatted string
    def process_and_output(time_unit, max_process, waiting_process):
        max_process.instructions -= 1
        print(f"Time Unit: {time_unit}, PID {max_process.pid} executing, Instructions left: {max_process.instructions}")
        waiting_text = ""
        i = 0
        for my_process in waiting_process:
            i += 1
            if i % 6 == 0:
                waiting_text = f"{waiting_text}\n"
            waiting_time = time_unit - my_process.arrival_time - (my_process.initial_instructions - my_process.instructions) + 1
            waiting_text = f"{waiting_text}PID {my_process.pid} waiting time={waiting_time}. "
        if waiting_text != "":
            print(waiting_text)
 
#running modules
if __name__ == "__main__":
    my_scheduler = Scheduler()
    my_scheduler.load_processes()
    my_scheduler.run()